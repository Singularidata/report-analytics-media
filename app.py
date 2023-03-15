from datetime import datetime
import streamlit as st
import base64
import openpyxl

import pandas as pd
from processando_arquivos import processando_arquivo, processando_arquivo_e_tipo, dataframe_to_rows
from translators import translate_plataforma_from_google_ads, translate_plataforma_to_meta_ads, translate_campaign_type

st.title("ACOMP Report! Cultura Inglesa (CISP/CIISA)")

st.header("Arquivos CSV")
array_files = st.file_uploader(
    "Selecione todos os arquivos CSV para o relatório", type="csv", accept_multiple_files=True)

dados = {}
fonte_original = {
  'analytics': 'Google Analytics',
  'gads': 'Google Ads',
  'hubspot': 'Hubspot',
  'meta_ads': 'Meta Ads',
  'tiktok_ads': 'TikTok Ads'
}
translate_parse_dates = {
  'analytics': 'Data',
  'gads': 'Dia',
  'hubspot': 'Create Date',
  'meta_ads': 'Dia',
  'tiktok_ads': 'Date'
}
rename_columns = {}


#Crio um FOR. Assim, para cada arquivo enviado, o sistema calcula qual arquivo foi enviado e retorna para o usuário a informação
for file in array_files:
  file_content = file.read().decode('utf-8')
  try:
    file_content_corrected = processando_arquivo_e_tipo(file_content)
    tipo = file_content_corrected['tipo']
    dados[tipo] = pd.read_csv(file_content_corrected['arquivo'],  error_bad_lines=False, parse_dates=[translate_parse_dates[tipo]]) 
    st.header(f"✅ - {fonte_original[tipo]}")
  except:
    st.header(f"🚫 - {file.name} - O arquivo está com erro. Por favor revise!")

lista_de_fontes = fonte_original.keys()
lista_de_dados = dados.keys()
fontes_que_faltam = list(set(lista_de_fontes) - set(lista_de_dados))
if(len(fontes_que_faltam) > 0):
  st.write(f'Ainda falta arquivo para gerar o Export:')
  for i in fontes_que_faltam:
    st.header(f"⚠️ - {fonte_original[i]}")


if (len(fontes_que_faltam) == 0):
  #Default
  plataforma = ''
  colunas_default = ['Dia', 'Plataforma','Campanha', 'Tipo', 'Source', 'Investimento','Leads', 'CPL', 'Impressões', 'Cliques', 'CTR', 'CPC médio']
  dados_por_dia = {}

  def proccess_data(rename_columns, platform):
    dados[plataforma] = dados[plataforma].rename(columns=rename_columns)
    if platform == 'meta_ads':
      dados[platform]['Plataforma'] = dados[platform]['Campanha'].apply(translate_plataforma_to_meta_ads) 
    else:
      dados[platform]['Plataforma'] = dados[platform]['Campanha'].apply(translate_plataforma_from_google_ads)
    dados[platform]['Tipo'] = dados[platform]['Campanha'].apply(translate_campaign_type)
    dados[platform]["Source"] = platform.replace("_", " ").title()
    dados[platform] = dados[platform].fillna(0)

  for plataforma in dados:
    if plataforma == 'hubspot':
      rename_columns = {'Create Date': 'Dia', 'Original Source Drill-Down 1': 'Campanha', 'Lifecycle Stage': 'Tipo Lead', 'Email': 'Leads'}
      proccess_data(rename_columns=rename_columns, platform=plataforma)
      dados_por_dia[plataforma] = dados[plataforma].groupby([dados[plataforma]['Dia'].dt.date, 'Plataforma', 'Campanha', 'Tipo', 'Source', 'Tipo Lead']).agg({'Leads': 'count'}).reset_index()

    elif plataforma == 'analytics':
      rename_columns={'Data': 'Dia', 'ABlab - Submissão de formulário (Conclusões da meta 10)': 'Leads', 'ABlab - Submissão de formulário (Taxa de conversão da meta 10)': 'Taxa de conv.'}
      dados[plataforma]['Data'] = dados[plataforma]['Data'].dt.date
      proccess_data(rename_columns=rename_columns, platform=plataforma)
      dados_por_dia[plataforma] = dados[plataforma][['Dia', 'Plataforma', 'Campanha', 'Tipo', 'Source', 'Usuários', 'Sessões', 'Leads', 'Taxa de conv.']]

    elif plataforma == 'gads':
      rename_columns={'Custo': 'Investimento', 'Impr.': 'Impressões', 'Impr': 'Impressões', 'Conversões': 'Leads', 'Custo / conv.': 'CPL', 'Custo / conv': 'CPL'}
      dados[plataforma]['Dia'] = dados[plataforma]['Dia'].dt.date
      proccess_data(rename_columns=rename_columns, platform=plataforma)
      dados_por_dia[plataforma] = dados[plataforma][colunas_default]
      dados_por_dia[plataforma] = dados_por_dia[plataforma].astype({'Cliques': 'int', 'Leads': 'int', 'Impressões': 'int', 'Impressões': 'int','Investimento': 'float'})

    elif plataforma == 'meta_ads':
      rename_columns={
      'Valor gasto (BRL)': 'Investimento', 
      'Custo por cadastro': 'CPL', 
      'CTR (taxa de cliques no link)': 'CTR', 
      'CPC (custo por clique no link)': 'CPC médio', 
      'Cliques no link': 'Cliques', 
      'Nome da campanha': 'Campanha', 
      'Conversas por mensagem iniciadas': 'Leads (wpp)'
      }
      proccess_data(rename_columns=rename_columns, platform=plataforma)
      #Leads Qualificados CIRJ
      if 'Ablab Performance - CIISA' in dados[plataforma].columns:
        dados[plataforma]['Cadastros no site'] = dados[plataforma]['Ablab Performance - CIISA']
      if 'CISP - mkt' in dados[plataforma].columns:
        dados[plataforma]['Cadastros no site'] = dados[plataforma]['CISP - mkt']
      if not "Leads (wpp)" in dados[plataforma].columns:
        dados[plataforma]['Leads (wpp)'] = 0
      if not "Cadastros na Meta" in dados[plataforma].columns:
        dados[plataforma]['Cadastros na Meta'] = 0
      
      dados[plataforma]['Leads'] = dados[plataforma]['Cadastros no site'].astype(int) + dados[plataforma]['Leads (wpp)'].astype(int) + dados[plataforma]['Cadastros na Meta'].astype(int)
      dados[plataforma]['CPL'] = dados[plataforma]['Investimento'].div(dados[plataforma]['Leads']).round(2)
      dados_por_dia[plataforma] = dados[plataforma].groupby([dados[plataforma]['Dia'].dt.date, 'Plataforma', 'Campanha', 'Tipo', 'Source']).agg({'Investimento': 'sum', 'Leads': 'sum', 'CPL': 'mean', 'Impressões': 'sum', 'Cliques': 'sum', 'CTR': 'mean', 'CPC médio': 'mean', 'Leads (wpp)': 'sum'}).reset_index()
      dados_por_dia[plataforma] = dados_por_dia[plataforma].fillna(0)
      dados_por_dia[plataforma] = dados_por_dia[plataforma].astype({'Cliques': 'int', 'Leads': 'int', 'Impressões': 'int'})
      dados_por_dia[plataforma]['Taxa de conv.'] = dados_por_dia[plataforma]['Leads'].div(dados_por_dia[plataforma]['Cliques']).round(4)*100
      dados_por_dia[plataforma]['CPL'] = dados_por_dia[plataforma]['Investimento'].div(dados_por_dia[plataforma]['Leads']).round(2)
      dados_por_dia[plataforma]['CTR'] = dados_por_dia[plataforma][dados_por_dia[plataforma]['Plataforma'] == 'Meta Ads']['Cliques'].div(dados_por_dia[plataforma][dados_por_dia[plataforma]['Plataforma'] == 'Meta Ads']['Impressões']).round(4)*100
      dados_por_dia[plataforma] = dados_por_dia[plataforma].fillna(0)
      dados_por_dia[plataforma]['Taxa de conv.'] = dados_por_dia[plataforma]['Taxa de conv.'].round(2).map(str) +'%'
      dados_por_dia[plataforma]['CTR'] = dados_por_dia[plataforma]['CTR'].round(2).map(str) +'%'

    elif plataforma == 'tiktok_ads':
      rename_columns = {'Date': 'Dia', 'Campaign name': 'Campanha', 'Cost': 'Investimento', 'Total Submit Form': 'Leads', 'Impression': 'Impressões', 'Click': 'Cliques', 'CPC': 'CPC médio'}
      proccess_data(rename_columns=rename_columns, platform=plataforma)
      dados[plataforma]['Plataforma'] = dados[plataforma]['Plataforma'].str.replace('Outros', 'TikTok') 
      dados[plataforma]['CPL'] = dados[plataforma]['Investimento'].div(dados[plataforma]['Leads']).round(2)
      dados_por_dia[plataforma] = dados[plataforma].groupby([dados[plataforma]['Dia'].dt.date, 'Plataforma', 'Campanha', 'Tipo', 'Source']).agg({'Investimento': 'sum', 'Leads': 'sum', 'CPL': 'mean', 'Impressões': 'sum', 'Cliques': 'sum', 'CTR': 'mean', 'CPC médio': 'mean'}).reset_index()
      dados_por_dia[plataforma]['Taxa de conv.'] = dados_por_dia[plataforma]['Leads'].div(dados_por_dia[plataforma]['Cliques']).round(4)*100
      dados_por_dia[plataforma] = dados_por_dia[plataforma].fillna(0)
      dados_por_dia[plataforma]['Taxa de conv.'] = dados_por_dia[plataforma]['Taxa de conv.'].round(2).map(str) +'%'

  acumulado = pd.concat(dados_por_dia.values(), ignore_index=True)
  acumulado = acumulado[['Dia', 'Plataforma', 'Campanha',	'Tipo',	'Source',	'Usuários',	'Sessões',	'Leads',	'Taxa de conv.',	'Investimento',	'CPL',	'Impressões',	'Cliques',	'CTR',	'CPC médio',	'Leads (wpp)',	'Tipo Lead']]
  acumulado = acumulado.fillna(0)

  workbook =  openpyxl.load_workbook('./src/acomp_data_report.xlsx')

  workbook.remove(workbook.get_sheet_by_name('acumulado'))
  acumulado_worksheet = workbook.create_sheet('acumulado')
  for r in dataframe_to_rows(acumulado, index=False, header=True):
    acumulado_worksheet.append(r)
  
  for plat in dados_por_dia.keys():
    plat_raw = plat.replace("_", " ").replace('analytics', 'ga')
    workbook.remove(workbook.get_sheet_by_name(plat_raw))
    data_worksheet = workbook.create_sheet(plat_raw)
    for r in dataframe_to_rows(dados_por_dia[plat], index=False, header=True):
      data_worksheet.append(r)
  
  workbook.save('./acomp_data_report.xlsx')
  with open('./acomp_data_report.xlsx', 'rb') as f:
    excel_b64 = base64.b64encode(f.read())

  st.markdown(
      f'Agora você já pode baixar o Excel pronto clicando no link: <a href="data:application/octet-stream;base64,{excel_b64.decode()}" download="acomp_data_report.xlsx">acomp_data_report.xlsx</a>', unsafe_allow_html=True)

  st.button("Processar novamente!")
